import pandas as pd
import openpyxl


def scrape_worp_sheet(worksheet_path):
    df = pd.read_csv(worksheet_path)
    print(df.to_string())
    return df


def scrape_waldman_workbook(workbook_path):
    # data_only: discard excel formulas, only take computed values for sheet
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    nfl_teams = ["ARI", "ATL", "BALT", "BUF", "CAR", "CHI", "CIN", "CLE", "DAL", "DEN", "DET", "GB", "HOU", "IND",
                 "JAX", "KC", "LV", "LAC", "LAR", "MIA", "MIN", "NE", "NO", "NYG", "NYJ", "PHI", "PIT", "SF", "SEA",
                 "TB", "TEN", "WAS"]
    qb_rows = []
    rb_rows = []
    wr_rows = []
    te_rows = []

    # assumes that these headers are consistent throughout all 32 franchise worksheets
    qb_col_headers = ["Player", "Position", "Franchise", "Games",
                      "PAtt", "Comp", "Pct", "PYds", "PYds/Att", "PTDs", "PTDs/Att", "Int", "Int/Att",
                      "RuAtt", "RuYds", "RuYds/Att", "RuTDs", "RuTds/Att",
                      "Touch %", "Pct Pass", "Pct Run", "Fpts"]

    rb_col_headers = ["Player", "Position", "Franchise", "Games",
                      "RuAtt", "RuYds", "RuYds/Att", "RuTDs", "RuTDs/Att",
                      "Catch %", "Targets", "Rec Yds", "Yds/Rec", "TDs", "TDs/Rec",
                      "Touch %", "Pct Pass", "Pct Run", "Fpts", "PPR Fpts"]

    wr_col_headers = ["Player", "Position", "Franchise", "Games",
                      "Catch %", "Targets",	"Rec Yds", "Yds/Rec", "TDs", "TDs/Rec",
                      "RuAtt", "RuYds", "RuYds/Att", "RuTds", "RuTDs/Att",
                      "Touch %", "Pct Pass", "Pct Run", "Fpts", "PPR Fpts"]

    te_col_headers = ["Player", "Position", "Franchise", "Games",
                      "Catch %", "Targets",	"Rec Yds", "Yds/Rec", "TDs", "TDs/Rec",
                      "RuAtt", "RuYds", "RuYds/Att", "RuTDs", "RuTDs/Att",
                      "Touch %", "Pct Pass", "Pct Run", "Fpts", "PPR Fpts"]

    qb_rows += [qb_col_headers]
    rb_rows += [rb_col_headers]
    wr_rows += [wr_col_headers]
    te_rows += [te_col_headers]

    for franchise in nfl_teams:
        franchise_ws = wb[franchise]
        print("Processing " + franchise)
        # find year 1 projection start point
        RSP_YEAR_ONE_START = "RSP Projections: Year 1"
        RSP_YEAR_ONE_END = "RSP Projections: Year 2"
        # define data area for year 1
        start_row = None
        end_row = None
        count = 0
        for row in franchise_ws.iter_rows(min_row=1, min_col=1, max_col=1):
            count += 1
            for cell in row:
                if cell.value is not None:
                    if RSP_YEAR_ONE_START in str(cell.value):
                        print("Found Year 1 Projection start point!")
                        start_row = count + 1
                    if RSP_YEAR_ONE_END in str(cell.value):
                        print("Found Year 1 Projection start point!")
                        end_row = count - 1

        # iterate over sheet to find sections for each position
        qb_start = None
        qb_end = None
        qb_query = "QB"
        qbs_found = False
        rb_start = None
        rb_end = None
        rb_query = "RB"
        rbs_found = False
        wr_start = None
        wr_end = None
        wr_query = "WRs"
        wrs_found = False
        te_start = None
        te_end = None
        te_query = "Tes"
        tes_found = False
        end_query = "Total"
        row_count = 0

        for row in franchise_ws.iter_rows(min_row=start_row, min_col=1, max_row=end_row, max_col=1):
            if qbs_found and rbs_found and wrs_found and tes_found:
                # when all position groups are complete for the desired RSP year, move onto the next NFL franchise
                break
            for cell in row:
                if cell.value is not None:
                    # assumes order of data will always be QB, RB, WR, TE in that order...
                    if (not qb_start) and qb_query in str(cell.value):
                        print("Found QB start row: " + str(start_row+row_count+1))
                        qb_start = start_row + row_count + 1
                    if (not rb_start) and rb_query in str(cell.value):
                        print("Found RB start row: " + str(start_row+row_count+1))
                        rb_start = start_row + row_count + 1
                    if (not wr_start) and wr_query in str(cell.value):
                        print("Found WR start row: " + str(start_row+row_count+1))
                        wr_start = start_row + row_count + 1
                    if (not te_start) and te_query in str(cell.value):
                        print("Found TE start row: " + str(start_row+row_count+1))
                        te_start = start_row + row_count + 1
                    if end_query in str(cell.value):
                        # subtract one to stop before the positional totals
                        if not qbs_found:
                            qb_end = start_row + row_count - 1
                            print("Found QB end row: " + str(qb_end))
                            qbs_found = True
                        elif not rbs_found:
                            rb_end = start_row + row_count - 1
                            print("Found RB end row: " + str(rb_end))
                            rbs_found = True
                        elif not wrs_found:
                            wr_end = start_row + row_count - 1
                            print("Found WR end row: " + str(wr_end))
                            wrs_found = True
                        elif not tes_found:
                            te_end = start_row + row_count - 1
                            print("Found TE end row: " + str(te_end))
                            tes_found = True
                        else:
                            raise NotImplementedError("Error handling positional sections...")
            row_count += 1

        # aggregate entries in a data frame, to be merged with worp data frames
        for row in franchise_ws.iter_rows(min_row=qb_start, max_row=qb_end, values_only=True):
            if row[0] is None:
                continue
            qb = str(row[0])
            print(qb)
            qb_pos = qb.split("*")[0]
            qb_name = qb.split("*")[1].rsplit("-", 1)[0].strip()
            print("QB Name: " + qb_name)
            try:
                qb_last_name = qb_name.split(",")[0].strip()
                qb_first_name = qb_name.split(",")[1].strip()
            except IndexError as E:
                qb_last_name = qb_name.split(" ")[0].strip()
                qb_first_name = qb_name.split(" ")[1].strip()
            qb_ref_name = qb_first_name + " " + qb_last_name
            try:
                qb_team = qb.split("*")[1].split("-")[1].strip()
            except IndexError as E:
                qb_team = ""
            print("QB Name: " + qb_ref_name)
            print("QB Position: " + qb_pos)
            print("QB Team: " + qb_team)
            qb_row = [qb_ref_name, qb_pos, qb_team]
            try:
                # if data missing for beginning of row, just discard info for that player
                for i in range(1, 4):
                    qb_row.append(row[i])
                # remaining passing projections
                qb_row.append(row[3] / row[2])
                qb_row.append(row[5])
                qb_row.append(row[5] / row[2])
                qb_row.append(row[7])
                qb_row.append(row[7] / row[2])
                qb_row.append(row[9])
                qb_row.append(row[9] / row[2])
                # rushing projections
                qb_row.append(row[12])
                qb_row.append(row[13])
                qb_row.append(row[13] / row[12])
                qb_row.append(row[15])
                qb_row.append(row[15] / row[12])
                # touches
                qb_pass_att = franchise_ws.cell(row=qb_end + 1, column=3).value
                qb_rush_att = franchise_ws.cell(row=qb_end + 1, column=13).value
                touch_pct = (row[2] + row[12]) / (qb_rush_att + qb_pass_att)
                touch_pct_pass = row[2] / qb_pass_att
                touch_pct_rush = row[12] / qb_rush_att
                qb_yards = row[5]
                qb_pass_tds = row[7]
                qb_ints = row[9]
                qb_rush_yds = row[13]
                qb_rush_tds = row[15]
                qb_fantasy_score = (qb_yards * .04) + (qb_pass_tds * 6) - (qb_ints * 2) + (qb_rush_yds * .1) + (
                        qb_rush_tds * 6)
                qb_row.append(touch_pct)
                qb_row.append(touch_pct_pass)
                qb_row.append(touch_pct_rush)
                qb_row.append(qb_fantasy_score)
                qb_rows.append(qb_row)
            except:
                # skip this player if data is missing, throwing exceptions
                continue

        for row in franchise_ws.iter_rows(min_row=rb_start, max_row=rb_end, values_only=True):
            if row[0] is None:
                continue
            rb = str(row[0]).strip()
            print(rb)
            rb_pos = rb.split("*")[0].strip()
            rb_name = rb.split("*")[1].rsplit("-", 1)[0].strip()
            print(rb_name)
            try:
                rb_last_name = rb_name.split(",")[0].strip()
                rb_first_name = rb_name.split(",")[1].strip()
            except IndexError as E:
                rb_last_name = rb_name.split(" ")[0].strip()
                rb_first_name = rb_name.split(" ")[1].strip()
            rb_ref_name = rb_first_name + " " + rb_last_name
            try:
                rb_team = rb.split("*")[1].split("-")[1].strip()
            except IndexError as E:
                rb_team = ""
            print("RB Name: " + rb_ref_name)
            print("RB Position: " + rb_pos)
            print("RB Team: " + rb_team)
            rb_row = [rb_ref_name, rb_pos, rb_team]
            try:
                rb_rush_atts = row[2]
                rb_rush_yds = row[3]
                rb_rush_yds_per_att = rb_rush_yds/rb_rush_atts
                rb_rush_tds = row[5]
                rb_rush_tds_per_att = rb_rush_tds/rb_rush_atts
                for value in row[1:4]:
                    rb_row.append(value)
                # remaining rushing stats
                rb_row.append(rb_rush_yds_per_att)
                rb_row.append(rb_rush_tds)
                rb_row.append(rb_rush_tds_per_att)
                # receiving stats
                rb_catch_pct = row[10]/row[9]
                rb_targets = row[9]
                rb_row.append(rb_catch_pct)
                rb_row.append(row[9])
                rb_row.append(row[10])
                rb_row.append(row[11])
                rb_row.append(row[11]/row[10])
                rb_row.append(row[12])
                rb_row.append(row[12]/row[10])
                # touches
                rb_room_rush_att = franchise_ws.cell(row=rb_end + 1, column=3).value
                rb_room_targets = franchise_ws.cell(row=rb_end + 1, column=10).value
                rb_touch_pct = (rb_rush_att + rb_targets)/(rb_room_rush_att+rb_room_targets+)
                rb_rows.append(rb_row)
            except:
                continue

        for row in franchise_ws.iter_rows(min_row=wr_start, max_row=wr_end):
            if row[0].value is None:
                continue
            wr = str(row[0].value).strip()
            print(wr)
            wr_pos = wr.split("*")[0].strip()
            wr_name = wr.split("*")[1].rsplit("-", 1)[0].strip()
            print(wr_name)
            try:
                wr_last_name = wr_name.split(",")[0].strip()
                wr_first_name = wr_name.split(",")[1].strip()
            except IndexError as E:
                wr_last_name = wr_name.split(" ")[0].strip()
                wr_first_name = wr_name.split(" ")[1].strip()
            wr_ref_name = wr_first_name + " " + wr_last_name
            try:
                wr_team = wr.split("*")[1].split("-")[1].strip()
            except IndexError as E:
                wr_team = ""
            print("WR Name: " + wr_ref_name)
            print("WR Position: " + wr_pos)
            print("WR Team: " + wr_team)
            wr_row = [wr_ref_name, wr_pos, wr_team]
            for cell in row[1:]:
                wr_row.append(cell.value)
            wr_rows.append(wr_row)

        for row in franchise_ws.iter_rows(min_row=te_start, max_row=te_end):
            if row[0].value is None:
                continue
            te = str(row[0].value).strip()
            print(te)
            te_pos = te.split("*")[0].strip()
            te_name = te.split("*")[1].rsplit("-", 1)[0].strip()
            print(te_name)
            try:
                te_last_name = te_name.split(",")[0].strip()
                te_first_name = te_name.split(",")[1].strip()
            except IndexError as E:
                te_last_name = te_name.split(" ")[0].strip()
                te_first_name = te_name.split(" ")[1].strip()
            te_ref_name = te_first_name + " " + te_last_name
            try:
                te_team = te.split("*")[1].split("-")[1].strip()
            except IndexError as E:
                te_team = ""
            print("TE Name: " + te_ref_name)
            print("TE Position: " + te_pos)
            print("TE Team: " + te_team)
            te_row = [te_ref_name, te_pos, te_team]
            for cell in row[1:]:
                te_row.append(cell.value)
            te_rows.append(te_row)

    qb_df = pd.DataFrame(qb_rows)
    rb_df = pd.DataFrame(rb_rows)
    wr_df = pd.DataFrame(wr_rows)
    te_df = pd.DataFrame(te_rows)

    return qb_df, rb_df, wr_df, te_df


if __name__ == '__main__':
    worksheet_path = 'SFC Bushwood_WoRP_Table_2022.csv'
    waldman_wb = "June_RSP_Projections.xlsx"
    worp_df = scrape_worp_sheet(worksheet_path)
    qb_df, rb_df, wr_df, te_df = scrape_waldman_workbook(waldman_wb)

